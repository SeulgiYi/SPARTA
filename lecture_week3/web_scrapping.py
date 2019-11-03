from pprint import pprint
import requests
from bs4 import BeautifulSoup

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)
soup = BeautifulSoup(data.text, 'html.parser')

movies = soup.select('#old_content > table > tbody > tr')

from openpyxl import load_workbook
work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

temp_text = work_sheet.cell(row = 1, column = 1).value
print(temp_text)

if not('movie_ranking' in work_book.sheetnames):
    work_sheet2 = work_book.create_sheet('movie_ranking')
work_sheet2.cell(row = 1, column = 1, value='번호')
work_sheet2.cell(row = 1, column = 2, value='영화제목')
work_sheet2.cell(row = 1, column = 3, value='별점')

irow = 1
for movie in movies:
    a_tag = movie.select_one('td.title > div > a')
    ranking = movie.select_one('td.point')
    if a_tag is not None:
        row = [ irow, a_tag.text, ranking.text ]
        irow += 1
        for j, col in enumerate(row):
            work_sheet2.cell(row = irow, column = j+1, value = col)

work_book.save('prac01.xlsx')